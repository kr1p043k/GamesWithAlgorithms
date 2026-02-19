import pygame
import random
import math
import heapq
import numpy as np
from collections import deque
from abc import ABC, abstractmethod
import openpyxl  # для записи Excel

# Константы
WIDTH, HEIGHT = 600, 600
CELL_SIZE = 40
GRID_SIZE = 15
WIN_WIDTH, WIN_HEIGHT = GRID_SIZE * CELL_SIZE, GRID_SIZE * CELL_SIZE
FPS = 10

# Цвета
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
GREEN = (0, 255, 0)
RED = (255, 0, 0)
BLUE = (0, 0, 255)
YELLOW = (255, 255, 0)
GRAY = (128, 128, 128)

# Направления
UP = (0, -1)
DOWN = (0, 1)
LEFT = (-1, 0)
RIGHT = (1, 0)
DIRECTIONS = [UP, DOWN, LEFT, RIGHT]

speed = FPS  # текущая скорость (для ручной регулировки)

class SnakeGame:
    def __init__(self):
        self.grid_size = GRID_SIZE
        self.reset()
        self.start_time = pygame.time.get_ticks()

    def reset(self):
        self.snake = [(7, 7), (6, 7), (5, 7)]
        self.direction = RIGHT
        self.score = 0
        self.apple = self._random_apple()
        self.game_over = False
        self.won = False
        self.start_time = pygame.time.get_ticks()

    def _random_apple(self):
        while True:
            apple = (random.randint(0, self.grid_size-1), random.randint(0, self.grid_size-1))
            if apple not in self.snake:
                return apple

    def change_direction(self, new_dir):
        if (new_dir[0] * -1, new_dir[1] * -1) != self.direction:
            self.direction = new_dir

    def move(self):
        if self.game_over:
            return

        head = self.snake[0]
        new_head = (head[0] + self.direction[0], head[1] + self.direction[1])

        if (new_head[0] < 0 or new_head[0] >= self.grid_size or
            new_head[1] < 0 or new_head[1] >= self.grid_size):
            self.game_over = True
            return

        if new_head == self.apple:
            self.snake.insert(0, new_head)
            self.score += 1
            if self.score >= 20:
                self.won = True
                self.game_over = True
            else:
                self.apple = self._random_apple()
        else:
            self.snake.insert(0, new_head)
            self.snake.pop()

        if self.snake[0] in self.snake[1:]:
            self.game_over = True

    def get_state(self):
        return {
            'snake': self.snake.copy(),
            'apple': self.apple,
            'direction': self.direction,
            'grid_size': self.grid_size
        }

    def draw(self, screen):
        screen.fill(BLACK)
        for x in range(0, WIN_WIDTH, CELL_SIZE):
            pygame.draw.line(screen, GRAY, (x, 0), (x, WIN_HEIGHT))
        for y in range(0, WIN_HEIGHT, CELL_SIZE):
            pygame.draw.line(screen, GRAY, (0, y), (WIN_WIDTH, y))

        for i, segment in enumerate(self.snake):
            color = GREEN if i == 0 else (0, 200, 0)
            rect = pygame.Rect(segment[0]*CELL_SIZE, segment[1]*CELL_SIZE, CELL_SIZE, CELL_SIZE)
            pygame.draw.rect(screen, color, rect)
            pygame.draw.rect(screen, BLACK, rect, 2)

        rect = pygame.Rect(self.apple[0]*CELL_SIZE, self.apple[1]*CELL_SIZE, CELL_SIZE, CELL_SIZE)
        pygame.draw.rect(screen, RED, rect)
        pygame.draw.rect(screen, BLACK, rect, 2)

        font = pygame.font.Font(None, 36)
        text = font.render(f"Score: {self.score}", True, WHITE)
        screen.blit(text, (10, 10))
        if self.won:
            win_text = font.render("YOU WIN!", True, YELLOW)
            screen.blit(win_text, (WIN_WIDTH//2-50, WIN_HEIGHT//2))
        elif self.game_over:
            over_text = font.render("GAME OVER", True, RED)
            screen.blit(over_text, (WIN_WIDTH//2-70, WIN_HEIGHT//2))

        elapsed_ms = pygame.time.get_ticks() - self.start_time
        minutes = elapsed_ms // 60000
        seconds = (elapsed_ms % 60000) / 1000.0
        time_text = font.render(f"Time: {minutes:02d}:{seconds:06.4f}", True, WHITE)
        screen.blit(time_text, (WIN_WIDTH - 220, 10))

# ===================== АЛГОРИТМЫ ПОИСКА ПУТИ =====================

class PathFinder(ABC):
    def __init__(self, game_state):
        self.snake = game_state['snake']
        self.apple = game_state['apple']
        self.grid_size = game_state['grid_size']
        self.head = self.snake[0]
        self.obstacles = set(self.snake[1:])

    def is_valid(self, pos):
        x, y = pos
        return 0 <= x < self.grid_size and 0 <= y < self.grid_size and pos not in self.obstacles

    def get_neighbors(self, pos):
        x, y = pos
        neighbors = []
        for dx, dy in DIRECTIONS:
            nx, ny = x+dx, y+dy
            if self.is_valid((nx, ny)):
                neighbors.append((nx, ny))
        return neighbors

    def heuristic(self, pos):
        return abs(pos[0] - self.apple[0]) + abs(pos[1] - self.apple[1])

    @abstractmethod
    def get_direction(self):
        pass

class AStar(PathFinder):
    def get_direction(self):
        start = self.head
        goal = self.apple

        open_set = []
        closed_set = set()
        came_from = {}
        g_score = {start: 0}
        f_score = {start: self.heuristic(start)}

        heapq.heappush(open_set, (f_score[start], start))

        while open_set:
            current = heapq.heappop(open_set)[1]

            if current == goal:
                path = []
                while current in came_from:
                    path.append(current)
                    current = came_from[current]
                path.append(start)
                path.reverse()
                if len(path) > 1:
                    next_step = path[1]
                    return (next_step[0] - start[0], next_step[1] - start[1])
                return None

            closed_set.add(current)

            for neighbor in self.get_neighbors(current):
                if neighbor in closed_set:
                    continue
                tentative_g = g_score[current] + 1

                if neighbor not in g_score or tentative_g < g_score[neighbor]:
                    came_from[neighbor] = current
                    g_score[neighbor] = tentative_g
                    f_score[neighbor] = tentative_g + self.heuristic(neighbor)
                    heapq.heappush(open_set, (f_score[neighbor], neighbor))

        for dx, dy in DIRECTIONS:
            new_pos = (start[0]+dx, start[1]+dy)
            if self.is_valid(new_pos):
                return (dx, dy)
        return None

class ACO(PathFinder):
    """Улучшенный муравьиный алгоритм для поиска пути"""
    def __init__(self, game_state):
        super().__init__(game_state)
        self.n_ants = 20              # увеличено
        self.n_iterations = 30         # увеличено
        self.evaporation = 0.3          # уменьшено для сохранения феромона
        self.alpha = 1.0
        self.beta = 2.0
        self.Q = 70                   # константа для количества феромона
        self.pheromone = np.ones((self.grid_size, self.grid_size)) * 0.1

    def get_direction(self):
        best_path = None
        best_length = float('inf')

        for _ in range(self.n_iterations):
            paths = []          # только успешные пути
            lengths = []
            for _ in range(self.n_ants):
                path = self._construct_path()
                if path and path[-1] == self.apple:   # только пути, достигшие яблока
                    paths.append(path)
                    lengths.append(len(path))
                    if len(path) < best_length:
                        best_length = len(path)
                        best_path = path

            # Испарение
            self.pheromone *= (1 - self.evaporation)

            # Обновление феромонов от всех успешных муравьёв
            for path in paths:
                for pos in path:
                    self.pheromone[pos] += self.Q / len(path)

            # Элитизм: дополнительно усиливаем лучший путь
            if best_path:
                for pos in best_path:
                    self.pheromone[pos] += self.Q / best_length

        if best_path and len(best_path) > 1:
            next_step = best_path[1]
            return (next_step[0] - self.head[0], next_step[1] - self.head[1])

        # Fallback
        for dx, dy in DIRECTIONS:
            new_pos = (self.head[0]+dx, self.head[1]+dy)
            if self.is_valid(new_pos):
                return (dx, dy)
        return None

    def _construct_path(self):
        path = [self.head]
        current = self.head
        visited = set(path)
        max_steps = 150   # увеличено
        for _ in range(max_steps):
            if current == self.apple:
                break
            neighbors = self.get_neighbors(current)
            if not neighbors:
                break
            probabilities = []
            for n in neighbors:
                if n in visited:
                    prob = 0
                else:
                    tau = self.pheromone[n] ** self.alpha
                    eta = (1.0 / (self.heuristic(n) + 1)) ** self.beta
                    prob = tau * eta
                probabilities.append(prob)
            total = sum(probabilities)
            if total == 0:
                # если все вероятности нулевые, выбираем случайно из соседей (кроме посещённых)
                valid = [n for n in neighbors if n not in visited]
                if not valid:
                    valid = neighbors
                if not valid:
                    break
                next_pos = random.choice(valid)
            else:
                probabilities = [p/total for p in probabilities]
                next_pos = neighbors[np.random.choice(len(neighbors), p=probabilities)]
            path.append(next_pos)
            visited.add(next_pos)
            current = next_pos
        return path

class ABC(PathFinder):
    def __init__(self, game_state):
        super().__init__(game_state)
        self.colony_size = 20
        self.max_iter = 10
        self.limit = 5
        self.population = self._initialize_population()

    def _initialize_population(self):
        pop = []
        for _ in range(self.colony_size):
            path = self._random_walk()
            pop.append({
                'path': path,
                'fitness': self._fitness(path),
                'trials': 0
            })
        return pop

    def _random_walk(self, max_steps=20):
        path = [self.head]
        current = self.head
        visited = set(path)
        for _ in range(max_steps):
            if current == self.apple:
                break
            neighbors = self.get_neighbors(current)
            valid = [n for n in neighbors if n not in visited]
            if not valid:
                valid = neighbors
            if not valid:
                break
            next_pos = random.choice(valid)
            path.append(next_pos)
            visited.add(next_pos)
            current = next_pos
        return path

    def _fitness(self, path):
        if not path:
            return float('inf')
        last = path[-1]
        if last == self.apple:
            return len(path)
        else:
            return len(path) + self.heuristic(last) * 2

    def _mutate(self, path):
        if len(path) < 3:
            return self._random_walk()
        idx = random.randint(1, len(path)-1)
        new_path = path[:idx]
        current = path[idx-1]
        visited = set(new_path)
        for _ in range(15 - len(new_path)):
            if current == self.apple:
                break
            neighbors = self.get_neighbors(current)
            valid = [n for n in neighbors if n not in visited]
            if not valid:
                break
            next_pos = random.choice(valid)
            new_path.append(next_pos)
            visited.add(next_pos)
            current = next_pos
        return new_path

    def get_direction(self):
        for iteration in range(self.max_iter):
            for i in range(self.colony_size):
                new_path = self._mutate(self.population[i]['path'])
                new_fitness = self._fitness(new_path)
                if new_fitness < self.population[i]['fitness']:
                    self.population[i]['path'] = new_path
                    self.population[i]['fitness'] = new_fitness
                    self.population[i]['trials'] = 0
                else:
                    self.population[i]['trials'] += 1

            total_fitness = sum(1/(1+b['fitness']) for b in self.population)
            if total_fitness == 0:
                break
            probs = [(1/(1+b['fitness']))/total_fitness for b in self.population]
            for _ in range(self.colony_size):
                i = np.random.choice(range(self.colony_size), p=probs)
                new_path = self._mutate(self.population[i]['path'])
                new_fitness = self._fitness(new_path)
                if new_fitness < self.population[i]['fitness']:
                    self.population[i]['path'] = new_path
                    self.population[i]['fitness'] = new_fitness
                    self.population[i]['trials'] = 0

            for i in range(self.colony_size):
                if self.population[i]['trials'] > self.limit:
                    self.population[i]['path'] = self._random_walk()
                    self.population[i]['fitness'] = self._fitness(self.population[i]['path'])
                    self.population[i]['trials'] = 0

        best = min(self.population, key=lambda x: x['fitness'])
        if best['path'] and len(best['path']) > 1:
            next_step = best['path'][1]
            return (next_step[0] - self.head[0], next_step[1] - self.head[1])
        for dx, dy in DIRECTIONS:
            new_pos = (self.head[0]+dx, self.head[1]+dy)
            if self.is_valid(new_pos):
                return (dx, dy)
        return None

class PSO(PathFinder):
    def __init__(self, game_state):
        super().__init__(game_state)
        self.swarm_size = 20
        self.max_iter = 15
        self.w = 0.5
        self.c1 = 1.5
        self.c2 = 1.5

        self.global_best_fitness = float('inf')
        self.global_best_path = None
        self.swarm = self._initialize_swarm()

    def _initialize_swarm(self):
        swarm = []
        for _ in range(self.swarm_size):
            path = self._random_walk()
            fitness = self._fitness(path)
            if fitness < self.global_best_fitness:
                self.global_best_fitness = fitness
                self.global_best_path = path.copy()
            swarm.append({
                'path': path,
                'best_path': path.copy(),
                'best_fitness': fitness,
                'velocity': 0
            })
        return swarm

    def _random_walk(self, max_steps=50):
        path = [self.head]
        current = self.head
        visited = set(path)
        for _ in range(max_steps):
            if current == self.apple:
                break
            neighbors = self.get_neighbors(current)
            valid = [n for n in neighbors if n not in visited]
            if not valid:
                valid = neighbors
            if not valid:
                break
            next_pos = random.choice(valid)
            path.append(next_pos)
            visited.add(next_pos)
            current = next_pos
        return path

    def _fitness(self, path):
        if not path:
            return float('inf')
        last = path[-1]
        if last == self.apple:
            return len(path)
        else:
            return len(path) + self.heuristic(last) * 2

    def _combine_paths(self, path1, path2):
        min_len = min(len(path1), len(path2))
        i = 0
        while i < min_len and path1[i] == path2[i]:
            i += 1
        new_path = path1[:i]
        current = new_path[-1] if new_path else self.head
        visited = set(new_path)
        for _ in range(50 - len(new_path)):
            if current == self.apple:
                break
            neighbors = self.get_neighbors(current)
            valid = [n for n in neighbors if n not in visited]
            if not valid:
                break
            next_pos = random.choice(valid)
            new_path.append(next_pos)
            visited.add(next_pos)
            current = next_pos
        return new_path

    def get_direction(self):
        for iteration in range(self.max_iter):
            for p in self.swarm:
                r1, r2 = random.random(), random.random()
                if r1 < self.w:
                    new_path = p['path']
                else:
                    if r2 < 0.5:
                        new_path = self._combine_paths(p['path'], p['best_path'])
                    else:
                        new_path = self._combine_paths(p['path'], self.global_best_path)

                new_fitness = self._fitness(new_path)

                if new_fitness < p['best_fitness']:
                    p['best_path'] = new_path.copy()
                    p['best_fitness'] = new_fitness
                p['path'] = new_path

                if new_fitness < self.global_best_fitness:
                    self.global_best_fitness = new_fitness
                    self.global_best_path = new_path.copy()

        if self.global_best_path and len(self.global_best_path) > 1:
            next_step = self.global_best_path[1]
            return (next_step[0] - self.head[0], next_step[1] - self.head[1])
        for dx, dy in DIRECTIONS:
            new_pos = (self.head[0]+dx, self.head[1]+dy)
            if self.is_valid(new_pos):
                return (dx, dy)
        return None

# ===================== БЕНЧМАРК (новая функция) =====================

def run_benchmark():
    print("Starting benchmark: collecting 100 successful games per algorithm (timeout 60s per game)...")
    algorithms = ['A*', 'ABC', 'PSO', 'ACO']
    results = {alg: {'times': [], 'failures': 0} for alg in algorithms}

    # Параметры алгоритмов (текущие настройки)
    params = {
        'A*': {},
        'ABC': {'colony_size': 20, 'max_iter': 10, 'limit': 5},
        'PSO': {'swarm_size': 20, 'max_iter': 15, 'w': 0.5, 'c1': 1.5, 'c2': 1.5},
        'ACO': {'n_ants': 30, 'n_iterations': 50, 'evaporation': 0.3, 'alpha': 1.0, 'beta': 2.0}
    }

    for alg in algorithms:
        print(f"Running {alg}...")
        successes = 0
        total_games = 0
        while successes < 100:
            total_games += 1
            game = SnakeGame()
            start_time = pygame.time.get_ticks()
            timeout = False

            while not game.game_over and not timeout:
                # Проверка тайм-аута
                if pygame.time.get_ticks() - start_time > 60000:  # 60 секунд
                    timeout = True
                    break

                state = game.get_state()
                if alg == 'A*':
                    finder = AStar(state)
                elif alg == 'ABC':
                    finder = ABC(state)
                elif alg == 'PSO':
                    finder = PSO(state)
                elif alg == 'ACO':
                    finder = ACO(state)
                direction = finder.get_direction()
                if direction:
                    game.change_direction(direction)
                game.move()

            if game.won:
                elapsed = (pygame.time.get_ticks() - start_time) / 1000.0
                results[alg]['times'].append(elapsed)
                successes += 1
                if successes % 10 == 0:
                    print(f"  {alg}: {successes}/100 successes (total games: {total_games}, failures: {results[alg]['failures']})")
            else:
                results[alg]['failures'] += 1

        print(f"{alg} done. Successes: 100, failures: {results[alg]['failures']}")

    # Формирование Excel-файла (как и ранее, но теперь ровно 100 строк времен для каждого)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    # Заголовки
    ws['A1'] = 'A*'
    ws['C1'] = 'ABC'
    ws['E1'] = 'PSO'
    ws['G1'] = 'ACO'
    ws['A2'] = 'Время'
    ws['C2'] = 'Время'
    ws['E2'] = 'Время'
    ws['G2'] = 'Время'

    # Данные по играм (100 строк)
    for i in range(100):
        row = i + 3
        # A*
        ws.cell(row=row, column=1, value=str(results['A*']['times'][i]).replace('.', ','))
        # ABC
        ws.cell(row=row, column=3, value=str(results['ABC']['times'][i]).replace('.', ','))
        # PSO
        ws.cell(row=row, column=5, value=str(results['PSO']['times'][i]).replace('.', ','))
        # ACO
        ws.cell(row=row, column=7, value=str(results['ACO']['times'][i]).replace('.', ','))

    # Статистика и параметры (начинаем со 103 строки)
    row = 103

    # A* (только количество неудач)
    ws.cell(row=row, column=1, value='Кол-во неудач попыток')
    ws.cell(row=row, column=2, value=results['A*']['failures'])

    # ABC
    ws.cell(row=row, column=3, value='Колония')
    ws.cell(row=row, column=4, value=params['ABC']['colony_size'])
    row += 1
    ws.cell(row=row, column=3, value='Итерации')
    ws.cell(row=row, column=4, value=params['ABC']['max_iter'])
    row += 1
    ws.cell(row=row, column=3, value='Разведчики')
    ws.cell(row=row, column=4, value=params['ABC']['limit'])
    row += 1
    ws.cell(row=row, column=3, value='Количество неудачных попыток')
    ws.cell(row=row, column=4, value=results['ABC']['failures'])

    # PSO
    row = 103
    ws.cell(row=row, column=5, value='Колония')
    ws.cell(row=row, column=6, value=params['PSO']['swarm_size'])
    row += 1
    ws.cell(row=row, column=5, value='Итерации')
    ws.cell(row=row, column=6, value=params['PSO']['max_iter'])
    row += 1
    ws.cell(row=row, column=5, value='Инерция')
    ws.cell(row=row, column=6, value=str(params['PSO']['w']).replace('.', ','))
    row += 1
    ws.cell(row=row, column=5, value='Коэф. локального лучшего')
    ws.cell(row=row, column=6, value=str(params['PSO']['c1']).replace('.', ','))
    row += 1
    ws.cell(row=row, column=5, value='Коэф.глобал. Лучшего')
    ws.cell(row=row, column=6, value=str(params['PSO']['c2']).replace('.', ','))
    row += 1
    ws.cell(row=row, column=5, value='Кол-во неудач. Попыток')
    ws.cell(row=row, column=6, value=results['PSO']['failures'])

    # ACO
    row = 103
    ws.cell(row=row, column=7, value='Кол-во Муравьев')
    ws.cell(row=row, column=8, value=params['ACO']['n_ants'])
    row += 1
    ws.cell(row=row, column=7, value='Итерации')
    ws.cell(row=row, column=8, value=params['ACO']['n_iterations'])
    row += 1
    ws.cell(row=row, column=7, value='Коэф. испарение')
    ws.cell(row=row, column=8, value=str(params['ACO']['evaporation']).replace('.', ','))
    row += 1
    ws.cell(row=row, column=7, value='Феромон')
    ws.cell(row=row, column=8, value=str(params['ACO']['alpha']).replace('.', ','))
    row += 1
    ws.cell(row=row, column=7, value='Эвристика')
    ws.cell(row=row, column=8, value=str(params['ACO']['beta']).replace('.', ','))
    row += 1
    ws.cell(row=row, column=7, value='Кол-во неудач. Попыток')
    ws.cell(row=row, column=8, value=results['ACO']['failures'])

    wb.save("snake_benchmark.xlsx")
    print("Benchmark completed. Results saved to snake_benchmark.xlsx")
# ===================== ОСНОВНОЙ ЦИКЛ =====================

def main():
    pygame.init()
    screen = pygame.display.set_mode((WIN_WIDTH, WIN_HEIGHT))
    pygame.display.set_caption("Snake Game with Auto-pilot")
    clock = pygame.time.Clock()

    game = SnakeGame()
    auto_mode = False
    algorithm = None
    algorithm_name = ""

    # Выбор режима при старте
    choosing = True
    font = pygame.font.Font(None, 36)
    while choosing:
        screen.fill(BLACK)
        text1 = font.render("Choose mode:", True, WHITE)
        text2 = font.render("1 - Manual", True, WHITE)
        text3 = font.render("2 - A*", True, WHITE)
        text4 = font.render("3 - ACO", True, WHITE)
        text5 = font.render("4 - ABC", True, WHITE)
        text6 = font.render("5 - PSO", True, WHITE)
        text7 = font.render("6 - Benchmark (100 runs each)", True, WHITE)  # новый пункт
        screen.blit(text1, (50, 50))
        screen.blit(text2, (50, 100))
        screen.blit(text3, (50, 150))
        screen.blit(text4, (50, 200))
        screen.blit(text5, (50, 250))
        screen.blit(text6, (50, 300))
        screen.blit(text7, (50, 350))
        pygame.display.flip()

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                return
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_1:
                    auto_mode = False
                    choosing = False
                elif event.key == pygame.K_2:
                    auto_mode = True
                    algorithm = "A*"
                    algorithm_name = "A*"
                    choosing = False
                elif event.key == pygame.K_3:
                    auto_mode = True
                    algorithm = "ACO"
                    algorithm_name = "ACO"
                    choosing = False
                elif event.key == pygame.K_4:
                    auto_mode = True
                    algorithm = "ABC"
                    algorithm_name = "ABC"
                    choosing = False
                elif event.key == pygame.K_5:
                    auto_mode = True
                    algorithm = "PSO"
                    algorithm_name = "PSO"
                    choosing = False
                elif event.key == pygame.K_6:
                    # Запускаем бенчмарк и завершаем программу
                    run_benchmark()
                    pygame.quit()
                    return

    # Основной игровой цикл (для ручного или автономного режима)
    running = True
    while running:
        clock.tick(FPS if not auto_mode else speed)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if not auto_mode and not game.game_over:
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_UP:
                        game.change_direction(UP)
                    elif event.key == pygame.K_DOWN:
                        game.change_direction(DOWN)
                    elif event.key == pygame.K_LEFT:
                        game.change_direction(LEFT)
                    elif event.key == pygame.K_RIGHT:
                        game.change_direction(RIGHT)

        if auto_mode and not game.game_over:
            state = game.get_state()
            if algorithm == "A*":
                finder = AStar(state)
            elif algorithm == "ACO":
                finder = ACO(state)
            elif algorithm == "ABC":
                finder = ABC(state)
            elif algorithm == "PSO":
                finder = PSO(state)
            else:
                finder = None

            if finder:
                direction = finder.get_direction()
                if direction:
                    game.change_direction(direction)

        game.move()
        game.draw(screen)

        if game.game_over:
            font = pygame.font.Font(None, 36)
            restart_text = font.render("Press SPACE to restart or ESC to quit", True, WHITE)
            screen.blit(restart_text, (50, WIN_HEIGHT-50))
            pygame.display.flip()
            waiting = True
            while waiting:
                for event in pygame.event.get():
                    if event.type == pygame.QUIT:
                        running = False
                        waiting = False
                    if event.type == pygame.KEYDOWN:
                        if event.key == pygame.K_SPACE:
                            game.reset()
                            waiting = False
                        elif event.key == pygame.K_ESCAPE:
                            running = False
                            waiting = False
        else:
            pygame.display.flip()

    pygame.quit()

if __name__ == "__main__":
    main()